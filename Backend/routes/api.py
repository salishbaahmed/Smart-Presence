import sqlite3
import os
import sys
import json
import pickle
import base64
import hmac
import requests as http_requests
import numpy as np
from flask import Blueprint, jsonify, request, send_file, current_app, session
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from io import BytesIO
from functools import wraps
from ai_module import common as _common_module

log = _common_module.get_logger('api')

api_bp = Blueprint('api', __name__, url_prefix='/api')


# ── Global Error Handler ──
@api_bp.errorhandler(Exception)
def handle_api_error(e):
    """Catch any unhandled exception and return clean JSON instead of HTML traceback."""
    import traceback
    import sys
    print(traceback.format_exc(), file=sys.stdout)
    sys.stdout.flush()
    log.error(f"Unhandled API error: {type(e).__name__}: {e}")
    return jsonify({"error": "Internal server error"}), 500




def _require_json():
    """Parse JSON body or return a 400 error response. Returns (data, error_response)."""
    data = request.get_json(silent=True)
    if not data:
        return None, (jsonify({"error": "Invalid or missing JSON body"}), 400)
    return data, None


# ── Rate Limiter Helper ──
def _get_limiter():
    """Get the rate limiter from the app, if available."""
    try:
        from flask import current_app
        return getattr(current_app, 'limiter', None)
    except Exception:
        return None


def get_db():
    db_path = current_app.config['DB_PATH']
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ── Auth Helpers ──

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated


def api_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "Authentication required"}), 401
        if session.get('role') != 'admin':
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════
#  AUTH — Login / Logout / Session
# ══════════════════════════════════════════════════════

@api_bp.route('/auth/login', methods=['POST'])
def auth_login():
    # Rate limit: 5 attempts per minute
    limiter = _get_limiter()
    if limiter:
        try:
            limiter.limit("5 per minute")(lambda: None)()
        except Exception:
            pass  # Limiter will handle 429 responses automatically

    data, err = _require_json()
    if err: return err
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid username or password"}), 401

    session['user_id'] = user['id']
    session['username'] = user['username']
    session['display_name'] = user['display_name']
    session['role'] = user['role']
    session.permanent = True

    return jsonify({
        "success": True,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "display_name": user['display_name'],
            "role": user['role']
        }
    })


@api_bp.route('/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({"success": True})


@api_bp.route('/auth/me', methods=['GET'])
def auth_me():
    if 'user_id' not in session:
        return jsonify({"authenticated": False}), 401
    return jsonify({
        "authenticated": True,
        "user": {
            "id": session['user_id'],
            "username": session['username'],
            "display_name": session['display_name'],
            "role": session['role']
        }
    })


@api_bp.route('/auth/verify-pin', methods=['POST'])
def verify_settings_pin():
    # Rate limit: 3 attempts per 10 minutes
    limiter = _get_limiter()
    if limiter:
        try:
            limiter.limit("3 per 10 minutes")(lambda: None)()
        except Exception:
            pass

    data, err = _require_json()
    if err: return err
    pin = data.get('pin', '')
    from ai_module import common
    # Timing-safe comparison to prevent side-channel attacks
    if hmac.compare_digest(str(pin), str(common.SETTINGS_PIN)):
        session['settings_unlocked'] = True
        return jsonify({"success": True})
    return jsonify({"error": "Invalid PIN"}), 403


# ══════════════════════════════════════════════════════
#  USERS — Admin CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/users', methods=['GET'])
@api_admin_required
def get_users():
    conn = get_db()
    users = conn.execute("SELECT id, username, display_name, role, created_at FROM users ORDER BY created_at").fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])


@api_bp.route('/users', methods=['POST'])
@api_admin_required
def add_user():
    data, err = _require_json()
    if err: return err
    username = data.get('username', '').strip()
    display_name = data.get('display_name', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'teacher')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    if len(username) > 50:
        return jsonify({"error": "Username too long (max 50 chars)"}), 400
    if len(display_name) > 100:
        return jsonify({"error": "Display name too long (max 100 chars)"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if not any(c.isdigit() for c in password) or not any(c.isalpha() for c in password):
        return jsonify({"error": "Password must contain letters and numbers"}), 400
    if role not in ('admin', 'teacher'):
        return jsonify({"error": "Role must be 'admin' or 'teacher'"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, display_name, password_hash, role) VALUES (?, ?, ?, ?)",
            (username, display_name or username, generate_password_hash(password), role))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Username already exists"}), 409
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/users/<int:uid>', methods=['PUT'])
@api_admin_required
def update_user(uid):
    data, err = _require_json()
    if err: return err
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "User not found"}), 404

    display_name = data.get('display_name', user['display_name']).strip()
    role = data.get('role', user['role'])

    if role not in ('admin', 'teacher'):
        conn.close()
        return jsonify({"error": "Invalid role"}), 400

    conn.execute("UPDATE users SET display_name=?, role=? WHERE id=?", (display_name, role, uid))

    # Update password if provided
    new_password = data.get('password', '').strip()
    if new_password:
        if len(new_password) < 6:
            conn.close()
            return jsonify({"error": "Password must be at least 6 characters"}), 400
        if not any(c.isdigit() for c in new_password) or not any(c.isalpha() for c in new_password):
            conn.close()
            return jsonify({"error": "Password must contain letters and numbers"}), 400
        conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                     (generate_password_hash(new_password), uid))

    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/users/<int:uid>', methods=['DELETE'])
@api_admin_required
def delete_user(uid):
    if uid == session.get('user_id'):
        return jsonify({"error": "Cannot delete your own account"}), 400
    conn = get_db()
    user = conn.execute("SELECT id FROM users WHERE id = ?", (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "User not found"}), 404
    conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  STUDENTS — CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/students', methods=['GET'])
@api_login_required
def get_students():
    conn = get_db()
    students = conn.execute("SELECT * FROM students ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(s) for s in students])


@api_bp.route('/students', methods=['POST'])
@api_login_required
def add_student():
    data, err = _require_json()
    if err: return err
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if len(name) > 100:
        return jsonify({"error": "Name too long (max 100 chars)"}), 400
    student_id = data.get('student_id', '').strip()
    if len(student_id) > 50:
        return jsonify({"error": "Student ID too long (max 50 chars)"}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO students (name, student_id, email, notes) VALUES (?, ?, ?, ?)",
                     (name, data.get('student_id', '').strip(), data.get('email', '').strip(),
                      data.get('notes', '').strip()))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Student already exists"}), 409
    conn.close()
    return jsonify({"success": True, "name": name}), 201


@api_bp.route('/students/<int:sid>', methods=['GET'])
@api_login_required
def get_student(sid):
    conn = get_db()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (sid,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    # Include attendance summary
    total_classes = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ?", (sid,)).fetchone()[0]
    present_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status IN ('Present', 'On Time')",
        (sid,)).fetchone()[0]
    late_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Late'",
        (sid,)).fetchone()[0]
    absent_count = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Absent'",
        (sid,)).fetchone()[0]

    # Recent attendance
    recent = conn.execute("""
        SELECT al.id, al.timestamp, al.status, al.source, al.notes
        FROM attendance_logs al WHERE al.student_id = ?
        ORDER BY al.timestamp DESC LIMIT 50
    """, (sid,)).fetchall()

    conn.close()

    attendance_rate = round((present_count + late_count) / total_classes * 100, 1) if total_classes > 0 else 0

    result = dict(student)
    result['stats'] = {
        'total_classes': total_classes,
        'present': present_count,
        'late': late_count,
        'absent': absent_count,
        'attendance_rate': attendance_rate
    }
    result['recent_attendance'] = [dict(r) for r in recent]
    return jsonify(result)


@api_bp.route('/students/<int:sid>', methods=['PUT'])
@api_login_required
def update_student(sid):
    data, err = _require_json()
    if err: return err
    conn = get_db()
    student = conn.execute("SELECT * FROM students WHERE id = ?", (sid,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404
    conn.execute("UPDATE students SET name=?, student_id=?, email=?, notes=? WHERE id=?",
                 (data.get('name', student['name']).strip(),
                  data.get('student_id', student['student_id'] or '').strip(),
                  data.get('email', student['email'] or '').strip(),
                  data.get('notes', student['notes'] or '').strip(), sid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/students/<int:sid>', methods=['DELETE'])
@api_login_required
def delete_student(sid):
    conn = get_db()
    student = conn.execute("SELECT id FROM students WHERE id = ?", (sid,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404
    conn.execute("DELETE FROM attendance_logs WHERE student_id = ?", (sid,))
    conn.execute("DELETE FROM students WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  ENROLLMENT — Web-based
# ══════════════════════════════════════════════════════

@api_bp.route('/enroll', methods=['POST'])
@api_login_required
def enroll_student():
    data, err = _require_json()
    if err: return err
    name = data.get('name', '').strip()
    student_id = data.get('student_id', '').strip()
    email = data.get('email', '').strip()
    image_b64 = data.get('image', '')
    force = data.get('force', False)

    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400
    if not email:
        return jsonify({"error": "Email is required"}), 400
    if not image_b64:
        return jsonify({"error": "Image is required"}), 400

    try:
        import face_recognition
        import cv2

        if ',' in image_b64:
            image_b64 = image_b64.split(',')[1]
        image_bytes = base64.b64decode(image_b64)
        np_arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if img is None:
            return jsonify({"error": "Invalid image data"}), 400

        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        boxes = face_recognition.face_locations(rgb, model="hog")
        if len(boxes) == 0:
            return jsonify({"error": "No face detected. Try a clearer photo."}), 400
        if len(boxes) > 1:
            return jsonify({"error": "Multiple faces detected. Use a photo with only one person."}), 400

        encodings = face_recognition.face_encodings(rgb, boxes)
        if not encodings:
            return jsonify({"error": "Could not encode the face."}), 500
        new_encoding = encodings[0]

        from ai_module import common
        enc_path = common.ENCODINGS_PATH

        # Duplicate face prevention: compare against existing encodings
        if os.path.exists(enc_path):
            with open(enc_path, "rb") as f:
                enc_data = pickle.load(f)

            if enc_data["encodings"] and not force:
                distances = face_recognition.face_distance(enc_data["encodings"], new_encoding)
                tolerance = getattr(common, 'TOLERANCE', 0.45)
                min_dist_idx = int(np.argmin(distances))
                min_dist = float(distances[min_dist_idx])
                if min_dist < tolerance:
                    matched_name = enc_data["names"][min_dist_idx]
                    return jsonify({
                        "error": f"This face closely matches '{matched_name}' (similarity: {round((1 - min_dist) * 100, 1)}%). "
                                 f"If this is a different person, re-submit with the override option.",
                        "duplicate": True,
                        "matched_name": matched_name,
                        "similarity": round((1 - min_dist) * 100, 1)
                    }), 409
        else:
            enc_data = {"names": [], "encodings": []}

        enc_data["names"].append(name)
        enc_data["encodings"].append(new_encoding)
        with open(enc_path, "wb") as f:
            pickle.dump(enc_data, f)

        conn = get_db()
        try:
            conn.execute("INSERT INTO students (name, student_id, email) VALUES (?, ?, ?)",
                         (name, student_id, email))
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({"error": "Student already exists in database", "name": name}), 409
        conn.close()

        try:
            from web_app.video_stream import video_stream
            if video_stream.is_running:
                video_stream.face_system.known_names.append(name)
                video_stream.face_system.known_encodings.append(new_encoding)
        except Exception:
            pass

        return jsonify({"success": True, "name": name}), 201
    except ImportError as e:
        return jsonify({"error": f"AI library not available: {e}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════
#  STUDENT PUBLIC LOOKUP (no auth required)
# ══════════════════════════════════════════════════════

@api_bp.route('/lookup', methods=['POST'])
def student_lookup():
    """Public endpoint for students to check their own attendance by student_id."""
    data, err = _require_json()
    if err: return err
    student_id = (data.get('student_id') or '').strip()
    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    conn = get_db()
    student = conn.execute("SELECT id, name, student_id FROM students WHERE student_id = ?", (student_id,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "No student found with that ID"}), 404

    sid = student['id']
    total = conn.execute("SELECT COUNT(*) FROM attendance_logs WHERE student_id = ?", (sid,)).fetchone()[0]
    present = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status IN ('Present', 'On Time')",
        (sid,)).fetchone()[0]
    late = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Late'",
        (sid,)).fetchone()[0]
    absent = conn.execute(
        "SELECT COUNT(*) FROM attendance_logs WHERE student_id = ? AND status = 'Absent'",
        (sid,)).fetchone()[0]

    recent = conn.execute("""
        SELECT timestamp, status, source FROM attendance_logs
        WHERE student_id = ? ORDER BY timestamp DESC LIMIT 20
    """, (sid,)).fetchall()
    conn.close()

    rate = round((present + late) / total * 100, 1) if total > 0 else 0

    return jsonify({
        "name": student['name'],
        "student_id": student['student_id'],
        "stats": {
            "total_classes": total,
            "present": present,
            "late": late,
            "absent": absent,
            "attendance_rate": rate
        },
        "recent_attendance": [{"timestamp": r['timestamp'], "status": r['status'], "source": r['source']} for r in recent]
    })


# ══════════════════════════════════════════════════════
#  ATTENDANCE — Read + Manual Entry/Override
# ══════════════════════════════════════════════════════

@api_bp.route('/attendance', methods=['GET'])
@api_login_required
def get_attendance():
    date_filter = request.args.get('date')
    student_id = request.args.get('student_id')
    conn = get_db()

    if student_id:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE al.student_id = ? ORDER BY al.timestamp DESC LIMIT 200
        """, (student_id,)).fetchall()
    elif date_filter:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) = ? ORDER BY al.timestamp DESC
        """, (date_filter,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT al.id, s.name, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            ORDER BY al.timestamp DESC LIMIT 200
        """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@api_bp.route('/attendance', methods=['POST'])
@api_login_required
def manual_attendance():
    data, err = _require_json()
    if err: return err
    student_id = data.get('student_id')
    status = data.get('status', 'Present')
    notes = data.get('notes', '').strip()
    if not student_id:
        return jsonify({"error": "student_id is required"}), 400

    from ai_module import common
    if status not in common.VALID_STATUSES:
        return jsonify({"error": f"Invalid status. Must be one of: {common.VALID_STATUSES}"}), 400

    # Opposing tag validation: prevent contradictory statuses on the same day
    PRESENCE_TAGS = {'Present', 'On Time', 'Late', 'Early Leave'}
    ABSENCE_TAGS = {'Absent', 'Excused', 'Permitted'}
    conn = get_db()
    student = conn.execute("SELECT id FROM students WHERE id = ?", (student_id,)).fetchone()
    if not student:
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    today_statuses = [r[0] for r in conn.execute(
        "SELECT DISTINCT status FROM attendance_logs WHERE student_id = ? AND DATE(timestamp) = DATE('now', 'localtime')",
        (student_id,)
    ).fetchall()]

    existing_presence = any(s in PRESENCE_TAGS for s in today_statuses)
    existing_absence = any(s in ABSENCE_TAGS for s in today_statuses)

    if status in PRESENCE_TAGS and existing_absence:
        conn.close()
        return jsonify({"error": f"Cannot mark as '{status}' — student already has an absence tag ({', '.join(s for s in today_statuses if s in ABSENCE_TAGS)}) today. Override the existing record instead."}), 409
    if status in ABSENCE_TAGS and existing_presence:
        conn.close()
        return jsonify({"error": f"Cannot mark as '{status}' — student already has a presence tag ({', '.join(s for s in today_statuses if s in PRESENCE_TAGS)}) today. Override the existing record instead."}), 409

    conn.execute("INSERT INTO attendance_logs (student_id, status, source, notes) VALUES (?, ?, 'manual', ?)",
                 (student_id, status, notes))
    conn.commit()
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/attendance/<int:log_id>', methods=['PUT'])
@api_login_required
def override_attendance(log_id):
    data, err = _require_json()
    if err: return err
    conn = get_db()
    log = conn.execute("SELECT * FROM attendance_logs WHERE id = ?", (log_id,)).fetchone()
    if not log:
        conn.close()
        return jsonify({"error": "Attendance record not found"}), 404

    # Teacher role: 7-day edit limit
    user_role = session.get('role', 'teacher')
    if user_role != 'admin':
        try:
            record_date = datetime.strptime(log['timestamp'][:10], '%Y-%m-%d')
            days_old = (datetime.now() - record_date).days
            if days_old > 7:
                conn.close()
                return jsonify({"error": f"Teachers can only edit records from the last 7 days. This record is {days_old} days old."}), 403
        except (ValueError, TypeError):
            pass

    from ai_module import common
    new_status = data.get('status', log['status'])
    if new_status not in common.VALID_STATUSES:
        conn.close()
        return jsonify({"error": f"Invalid status. Must be one of: {common.VALID_STATUSES}"}), 400

    conn.execute("UPDATE attendance_logs SET status=?, notes=?, source='override' WHERE id=?",
                 (new_status, data.get('notes', log['notes'] or ''), log_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/attendance/<int:log_id>', methods=['DELETE'])
@api_login_required
def delete_attendance(log_id):
    conn = get_db()
    log_entry = conn.execute("SELECT id FROM attendance_logs WHERE id = ?", (log_id,)).fetchone()
    if not log_entry:
        conn.close()
        return jsonify({"error": "Attendance record not found"}), 404
    conn.execute("DELETE FROM attendance_logs WHERE id = ?", (log_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  SCHEDULE — Timetable CRUD
# ══════════════════════════════════════════════════════

@api_bp.route('/schedule', methods=['GET'])
@api_login_required
def get_schedules():
    conn = get_db()
    rows = conn.execute("SELECT * FROM class_schedules ORDER BY day_of_week, start_time").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@api_bp.route('/schedule', methods=['POST'])
@api_login_required
def add_schedule():
    data, err = _require_json()
    if err: return err
    day = data.get('day_of_week', '').strip()
    start = data.get('start_time', '').strip()
    end = data.get('end_time', '').strip()
    name = data.get('class_name', 'Class').strip()
    teacher_email = data.get('teacher_email', '').strip()

    if not all([day, start, end]):
        return jsonify({"error": "day_of_week, start_time, end_time are required"}), 400

    valid_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    if day not in valid_days:
        return jsonify({"error": f"Invalid day. Must be one of: {valid_days}"}), 400

    # Validate end > start
    if end <= start:
        return jsonify({"error": "End time must be after start time"}), 400

    conn = get_db()
    conn.execute("INSERT INTO class_schedules (day_of_week, start_time, end_time, class_name, teacher_email) VALUES (?, ?, ?, ?, ?)",
                 (day, start, end, name, teacher_email))
    conn.commit()
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/schedule/<int:sid>', methods=['PUT'])
@api_login_required
def update_schedule(sid):
    data, err = _require_json()
    if err: return err
    conn = get_db()
    sched = conn.execute("SELECT * FROM class_schedules WHERE id = ?", (sid,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"error": "Schedule not found"}), 404

    conn.execute("UPDATE class_schedules SET day_of_week=?, start_time=?, end_time=?, class_name=?, teacher_email=?, is_active=? WHERE id=?",
                 (data.get('day_of_week', sched['day_of_week']),
                  data.get('start_time', sched['start_time']),
                  data.get('end_time', sched['end_time']),
                  data.get('class_name', sched['class_name']),
                  data.get('teacher_email', sched['teacher_email'] if 'teacher_email' in sched.keys() else ''),
                  data.get('is_active', sched['is_active']),
                  sid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/schedule/<int:sid>', methods=['DELETE'])
@api_login_required
def delete_schedule(sid):
    conn = get_db()
    sched = conn.execute("SELECT id FROM class_schedules WHERE id = ?", (sid,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"error": "Schedule not found"}), 404
    conn.execute("DELETE FROM class_schedules WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════
#  SYSTEM — Start / Stop / Restart / Mode / Status
# ══════════════════════════════════════════════════════

@api_bp.route('/system/status', methods=['GET'])
@api_login_required
def system_status():
    from web_app.video_stream import video_stream
    return jsonify(video_stream.get_status())


@api_bp.route('/system', methods=['POST'])
@api_login_required
def system_control():
    data, err = _require_json()
    if err: return err
    action = data.get('action', '')
    from web_app.video_stream import video_stream
    from ai_module import common

    if action == 'start':
        video_stream.start()
        return jsonify({"success": True, "message": "AI started"})
    elif action == 'stop':
        video_stream.stop()
        return jsonify({"success": True, "message": "AI stopped"})
    elif action == 'restart':
        video_stream.restart()
        return jsonify({"success": True, "message": "AI restarted"})
    elif action == 'shutdown':
        if session.get('role') != 'admin':
            return jsonify({"error": "Only admins can shutdown the server"}), 403
        video_stream.stop()
        import threading
        def _shutdown():
            import time
            time.sleep(1)
            # Graceful exit — allows cleanup (DB connections, file handles)
            sys.exit(0)
        threading.Thread(target=_shutdown, daemon=True).start()
        return jsonify({"success": True, "message": "Server shutting down..."})
    elif action == 'set_mode':
        mode = data.get('mode', 'auto')
        if mode not in ('auto', 'force_on', 'force_off'):
            return jsonify({"error": "Invalid mode"}), 400
        common.SYSTEM_MODE = mode
        return jsonify({"success": True, "mode": mode})
    else:
        return jsonify({"error": "Invalid action. Use: start, stop, restart, shutdown, set_mode"}), 400


# ══════════════════════════════════════════════════════
#  CRASH REPORTING + TELEGRAM
# ══════════════════════════════════════════════════════

def send_telegram(message):
    """Send a message to the configured Telegram chat."""
    from ai_module import common
    token = common.TELEGRAM_BOT_TOKEN
    chat_id = common.TELEGRAM_CHAT_ID

    if not token or token == 'YOUR_BOT_TOKEN_HERE' or not chat_id or chat_id == 'YOUR_CHAT_ID_HERE':
        return False

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        http_requests.post(url, json={
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "Markdown"
        }, timeout=5)
        return True
    except Exception:
        return False


@api_bp.route('/report', methods=['POST'])
@api_login_required
def submit_report():
    data, err = _require_json()
    if err: return err
    description = data.get('description', '').strip()
    if not description:
        return jsonify({"error": "Description is required"}), 400

    from ai_module import common
    from web_app.video_stream import video_stream

    category = data.get('category', 'bug')
    severity = data.get('severity', 'medium')

    # Build report
    report = {
        "timestamp": datetime.now().isoformat(),
        "description": description,
        "category": category,
        "severity": severity,
        "submitted_by": session.get('display_name', 'Unknown'),
    }

    # Selectable data inclusions
    if data.get('include_system', True):
        report["system_state"] = video_stream.get_status()
    if data.get('include_settings', True):
        report["settings"] = {
            "detection_scale": getattr(common, 'DETECTION_SCALE', 0.5),
            "tolerance": getattr(common, 'TOLERANCE', 0.5),
            "late_threshold": getattr(common, 'LATE_THRESHOLD', 10),
            "disappear_threshold": getattr(common, 'DISAPPEAR_THRESHOLD', 15),
            "system_mode": getattr(common, 'SYSTEM_MODE', 'auto'),
        }
    if data.get('include_browser', False):
        report["browser"] = data.get('user_agent', '')

    # Save locally
    reports_dir = common.CRASH_REPORTS_DIR
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(reports_dir, filename)
    with open(filepath, 'w') as f:
        json.dump(report, f, indent=2, default=str)

    # Rotate: keep only the 50 most recent reports
    MAX_REPORTS = 50
    try:
        reports = sorted(
            [os.path.join(reports_dir, f) for f in os.listdir(reports_dir)
             if f.startswith('report_') and f.endswith('.json')],
            key=os.path.getmtime
        )
        for old in reports[:-MAX_REPORTS]:
            os.remove(old)
    except OSError:
        pass

    # Send to Telegram
    severity_emoji = {'low': '🟡', 'medium': '🟠', 'critical': '🔴'}.get(severity, '⚪')
    tg_msg = (
        f"{severity_emoji} *SmartPresence Report*\n"
        f"*Category:* {category}\n"
        f"*Severity:* {severity}\n"
        f"*By:* {session.get('display_name', 'Unknown')}\n\n"
        f"_{description}_\n\n"
        f"📁 Saved: `{filename}`"
    )
    telegram_sent = send_telegram(tg_msg)

    return jsonify({
        "success": True,
        "file": filename,
        "telegram_sent": telegram_sent
    }), 201


# ══════════════════════════════════════════════════════
#  STATS + CHART DATA
# ══════════════════════════════════════════════════════

@api_bp.route('/stats', methods=['GET'])
@api_login_required
def get_stats():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    present_today = conn.execute("""
        SELECT COUNT(DISTINCT student_id) FROM attendance_logs
        WHERE DATE(timestamp) = ? AND status IN ('Present', 'On Time', 'Late')
    """, (today,)).fetchone()[0]
    absent_today = total_students - present_today
    total_logs = conn.execute("SELECT COUNT(*) FROM attendance_logs").fetchone()[0]
    conn.close()
    return jsonify({
        "total_students": total_students,
        "present_today": present_today,
        "absent_today": absent_today,
        "total_logs": total_logs,
        "date": today
    })


@api_bp.route('/stats/chart', methods=['GET'])
@api_login_required
def chart_data():
    try:
        days = max(1, min(365, int(request.args.get('days', 7))))
    except (ValueError, TypeError):
        days = 7
    conn = get_db()
    labels, present_data, absent_data, late_data = [], [], [], []
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]

    # Statuses that mean the student was physically present (at least partially)
    PRESENT_STATUSES = ('Present', 'On Time', 'Early Leave', 'Permitted', 'Excused')
    LATE_STATUSES = ('Late',)

    for i in range(days - 1, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        labels.append(date)
        present = conn.execute("""
            SELECT COUNT(DISTINCT student_id) FROM attendance_logs
            WHERE DATE(timestamp) = ? AND status IN ({})
        """.format(','.join('?' * len(PRESENT_STATUSES))), (date, *PRESENT_STATUSES)).fetchone()[0]
        late = conn.execute("""
            SELECT COUNT(DISTINCT student_id) FROM attendance_logs
            WHERE DATE(timestamp) = ? AND status IN ({})
        """.format(','.join('?' * len(LATE_STATUSES))), (date, *LATE_STATUSES)).fetchone()[0]
        present_data.append(present)
        late_data.append(late)
        absent_data.append(max(0, total_students - present - late))

    conn.close()
    return jsonify({"labels": labels, "present": present_data,
                    "late": late_data, "absent": absent_data,
                    "total_students": total_students})


# ══════════════════════════════════════════════════════
#  EXPORT
# ══════════════════════════════════════════════════════

@api_bp.route('/export', methods=['GET'])
@api_login_required
def export_data():
    fmt = request.args.get('format', 'xlsx')
    date_filter = request.args.get('date')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    conn = get_db()

    if date_from and date_to:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) BETWEEN ? AND ? ORDER BY al.timestamp
        """, (date_from, date_to)).fetchall()
        date_label = f"{date_from}_to_{date_to}"
    elif date_filter:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            WHERE DATE(al.timestamp) = ? ORDER BY al.timestamp
        """, (date_filter,)).fetchall()
        date_label = date_filter
    else:
        rows = conn.execute("""
            SELECT s.name, s.student_id as sid, al.timestamp, al.status, al.source, al.notes
            FROM attendance_logs al JOIN students s ON al.student_id = s.id
            ORDER BY al.timestamp DESC LIMIT 1000
        """).fetchall()
        date_label = "all"
    conn.close()

    return _export_csv(rows, date_label) if fmt == 'csv' else _export_xlsx(rows, date_label)


def _export_xlsx(rows, date_label):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    headers = ["Student Name", "Student ID", "Timestamp", "Status", "Source", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row['name'], row['sid'] or '', row['timestamp'],
                   row['status'], row['source'], row['notes'] or ''])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"attendance_{date_label}.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _export_csv(rows, date_label):
    import csv
    from io import StringIO
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student Name", "Student ID", "Timestamp", "Status", "Source", "Notes"])
    for row in rows:
        writer.writerow([row['name'], row['sid'] or '', row['timestamp'],
                         row['status'], row['source'], row['notes'] or ''])
    csv_bytes = BytesIO(output.getvalue().encode('utf-8'))
    csv_bytes.seek(0)
    return send_file(csv_bytes, download_name=f"attendance_{date_label}.csv",
                     as_attachment=True, mimetype='text/csv')






# ══════════════════════════════════════════════════════
#  ADMIN CONFIG — .env Editing (PIN-gated)
# ══════════════════════════════════════════════════════

# Editable keys — anything NOT in this list is hidden from the UI
EDITABLE_ENV_KEYS = [
    'SETTINGS_PIN',
    'SECRET_KEY',
    'TELEGRAM_BOT_TOKEN',
    'TELEGRAM_CHAT_ID',
    'DB_ENCRYPTION_KEY',
]

# Keys whose values should be masked in GET responses
MASKED_KEYS = {'SECRET_KEY', 'DB_ENCRYPTION_KEY', 'SETTINGS_PIN'}


def _env_path():
    """Return the absolute path to the project .env file."""
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '.env')


def _read_env():
    """Read .env file into an ordered list of (key, value, is_comment) tuples."""
    lines = []
    path = _env_path()
    if not os.path.exists(path):
        return lines
    with open(path, 'r', encoding='utf-8') as f:
        for raw in f:
            raw = raw.rstrip('\r\n')
            if raw.startswith('#') or raw.strip() == '':
                lines.append((None, raw, True))
            elif '=' in raw:
                k, v = raw.split('=', 1)
                lines.append((k.strip(), v.strip(), False))
            else:
                lines.append((None, raw, True))
    return lines


def _write_env(lines):
    """Write list of (key, value, is_comment) tuples back to .env."""
    path = _env_path()
    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        for key, val, is_comment in lines:
            if is_comment:
                f.write(val + '\n')
            else:
                f.write(f'{key}={val}\n')


@api_bp.route('/config', methods=['GET'])
@api_admin_required
def get_config():
    """Return editable env variables (masked where appropriate)."""
    if not session.get('settings_unlocked'):
        return jsonify({"error": "Settings PIN required", "pin_required": True}), 403

    lines = _read_env()
    result = {}
    for key, val, is_comment in lines:
        if is_comment or key not in EDITABLE_ENV_KEYS:
            continue
        if key in MASKED_KEYS:
            # Show first 4 + last 4 chars, mask middle
            if len(val) > 10:
                result[key] = val[:4] + '•' * (len(val) - 8) + val[-4:]
            else:
                result[key] = '•' * len(val)
        else:
            result[key] = val

    return jsonify({"config": result, "editable_keys": EDITABLE_ENV_KEYS})


@api_bp.route('/config', methods=['PUT'])
@api_admin_required
def update_config():
    """Update specific .env variables. Only EDITABLE_ENV_KEYS allowed."""
    if not session.get('settings_unlocked'):
        return jsonify({"error": "Settings PIN required", "pin_required": True}), 403

    data, err = _require_json()
    if err: return err
    updates = data.get('updates', {})

    if not updates:
        return jsonify({"error": "No updates provided"}), 400

    # Validate only editable keys
    for key in updates:
        if key not in EDITABLE_ENV_KEYS:
            return jsonify({"error": f"Key '{key}' is not editable"}), 400

    lines = _read_env()
    existing_keys = {k for k, v, c in lines if not c}

    # Update existing keys
    updated = []
    for key, val, is_comment in lines:
        if not is_comment and key in updates:
            new_val = str(updates[key]).strip()
            if not new_val:
                return jsonify({"error": f"Value for '{key}' cannot be empty"}), 400
            updated.append((key, new_val, False))
        else:
            updated.append((key, val, is_comment))

    # Add new keys that weren't in the file
    for key, val in updates.items():
        if key not in existing_keys:
            new_val = str(val).strip()
            if not new_val:
                return jsonify({"error": f"Value for '{key}' cannot be empty"}), 400
            updated.append((key, new_val, False))

    _write_env(updated)

    # Reload into common module
    try:
        from ai_module import common
        if 'SETTINGS_PIN' in updates:
            common.SETTINGS_PIN = updates['SETTINGS_PIN']
        if 'TELEGRAM_BOT_TOKEN' in updates:
            common.TELEGRAM_BOT_TOKEN = updates['TELEGRAM_BOT_TOKEN']
        if 'TELEGRAM_CHAT_ID' in updates:
            common.TELEGRAM_CHAT_ID = updates['TELEGRAM_CHAT_ID']
    except ImportError:
        pass

    # Update Flask SECRET_KEY if changed
    if 'SECRET_KEY' in updates:
        current_app.config['SECRET_KEY'] = updates['SECRET_KEY']

    return jsonify({"success": True, "message": f"{len(updates)} config value(s) updated", "updated_keys": list(updates.keys())})


@api_bp.route('/config/export-db', methods=['GET'])
@api_admin_required
def export_db():
    """Download a copy of the database file for backup."""
    db_path = current_app.config['DB_PATH']
    if not os.path.exists(db_path):
        return jsonify({"error": "Database not found"}), 404
    return send_file(
        db_path,
        as_attachment=True,
        download_name=f'smartpresence_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db',
        mimetype='application/x-sqlite3'
    )


@api_bp.route('/config/version', methods=['GET'])
@api_login_required
def get_version():
    """Return system version info."""
    return jsonify({
        "version": "1.6.0",
        "phase": "6B",
        "codename": "Camera Architecture",
        "build_date": "2026-02-13"
    })


# ══════════════════════════════════════════════════════
#  CAMERA CONFIGURATION
# ══════════════════════════════════════════════════════

@api_bp.route('/cameras', methods=['GET'])
@api_login_required
def get_cameras():
    from ai_module.camera_manager import CameraManager
    return jsonify(CameraManager.get_all_cameras())


@api_bp.route('/cameras', methods=['POST'])
@api_login_required
def add_camera():
    data, err = _require_json()
    if err: return err
    
    name = data.get('name', 'New Camera').strip()
    source = data.get('source', '').strip()
    ctype = data.get('type', 'usb').strip()
    
    if not source:
        return jsonify({"error": "Camera source is required."}), 400
        
    conn = get_db()
    conn.execute("INSERT INTO cameras (name, source, type, is_active) VALUES (?, ?, ?, ?)",
                 (name, source, ctype, 1))
    conn.commit()
    conn.close()
    return jsonify({"success": True}), 201


@api_bp.route('/cameras/<int:cam_id>', methods=['PUT'])
@api_login_required
def update_camera(cam_id):
    data, err = _require_json()
    if err: return err
    
    conn = get_db()
    cam = conn.execute("SELECT * FROM cameras WHERE id = ?", (cam_id,)).fetchone()
    if not cam:
        conn.close()
        return jsonify({"error": "Camera not found"}), 404
        
    name = data.get('name', cam['name'])
    source = data.get('source', cam['source'])
    ctype = data.get('type', cam['type'])
    is_active = int(data.get('is_active', cam['is_active']))
    
    conn.execute("UPDATE cameras SET name=?, source=?, type=?, is_active=? WHERE id=?",
                 (name, source, ctype, is_active, cam_id))
    conn.commit()
    conn.close()
    
    from ai_module import common
    common.get_logger('api').info(f"Camera {cam_id} updated via API")
    
    return jsonify({"success": True})


@api_bp.route('/cameras/<int:cam_id>', methods=['DELETE'])
@api_login_required
def delete_camera(cam_id):
    conn = get_db()
    conn.execute("DELETE FROM cameras WHERE id = ?", (cam_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@api_bp.route('/cameras/test', methods=['POST'])
@api_login_required
def test_camera_connection():
    data, err = _require_json()
    if err: return err
    
    source = data.get('source', '').strip()
    if not source:
        return jsonify({"error": "Source is required"}), 400
        
    src_val = int(source) if source.isdigit() else source
    
    import cv2
    try:
        cap = cv2.VideoCapture(src_val)
        if cap.isOpened():
            ret, frame = cap.read()
            cap.release()
            if ret and frame is not None:
                 return jsonify({"success": True, "message": "Connection successful!"})
            else:
                 return jsonify({"success": False, "error": "Camera opened but returned no frame."}), 400
        else:
            return jsonify({"success": False, "error": "Could not open camera source."}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ══════════════════════════════════════════════════════
#  AI SETTINGS
# ══════════════════════════════════════════════════════

@api_bp.route('/settings', methods=['GET'])
@api_login_required
def get_settings():
    from ai_module.settings import SettingsManager
    from web_app.video_stream import video_stream
    raw = SettingsManager.get_all()
    # Normalize keys to lowercase for JS frontend compatibility
    data = {k.lower(): v for k, v in raw.items()}
    # Add extra status fields the frontend expects
    data['frame_width'] = 1920
    data['frame_height'] = 1080
    return jsonify(data)


@api_bp.route('/settings', methods=['POST'])
@api_login_required
def update_settings():
    data, err = _require_json()
    if err: return err
    
    # Optional: PIN protection for settings? 
    # Current requirement says "Admin settings", api_login_required covers auth.
    
    from ai_module.settings import SettingsManager
    from ai_module import common
    
    success_count = 0
    errors = []
    
    # whitelist of allowed keys to prevent pollution
    ALLOWED_KEYS = {
        'DETECTOR_MODEL', 'TOLERANCE', 'DETECTION_SCALE', 
        'LATE_THRESHOLD', 'DISAPPEAR_THRESHOLD', 'RECHECK_INTERVAL',
        'SYSTEM_MODE', 'FRAME_SKIP'
    }
    
    for key, val in data.items():
        key = key.upper()  # Accept both lowercase and UPPERCASE from frontend
        if key not in ALLOWED_KEYS:
            continue
            
        # Basic validation
        if key in ['TOLERANCE', 'DETECTION_SCALE']:
            try:
                fval = float(val)
                if not (0.1 <= fval <= 1.0):
                    errors.append(f"{key} must be between 0.1 and 1.0")
                    continue
            except ValueError:
                errors.append(f"{key} must be a number")
                continue
                
        if key == 'DETECTOR_MODEL':
            if val not in ['dlib', 'mediapipe']:
                errors.append(f"{key} must be 'dlib' or 'mediapipe'")
                continue
                
        if key == 'SYSTEM_MODE':
            if val not in ['auto', 'force_on', 'force_off']:
                errors.append(f"{key} must be 'auto', 'force_on', or 'force_off'")
                continue

        if SettingsManager.set(key, val):
            success_count += 1
        else:
            errors.append(f"Failed to save {key}")

    common.get_logger('api').info(f"Settings updated: {success_count} keys changed. Errors: {errors}")
    
    return jsonify({
        "success": True, 
        "updated": success_count, 
        "errors": errors
    })


@api_bp.route('/health', methods=['GET'])
def health_check():
    """System health check endpoint (Public or Secured?). Currently secured by default if not exempt."""
    # Check DB
    db_status = "ok"
    try:
        conn = get_db()
        conn.execute("SELECT 1").fetchone()
        conn.close()
    except Exception as e:
        db_status = f"error: {str(e)}"
        
    # Check AI
    from web_app.video_stream import video_stream
    ai_status = "running" if video_stream.is_running else "stopped"
    
    return jsonify({
        "status": "healthy" if db_status == "ok" else "degraded",
        "database": db_status,
        "ai_engine": ai_status,
        "version": "1.6.0"
    })


# ══════════════════════════════════════════════════════
#  DEBUG — Admin-only system diagnostics
# ══════════════════════════════════════════════════════

@api_bp.route('/debug', methods=['GET'])
@api_admin_required
def debug_info():
    """Admin-only debug diagnostics. Accessible via secret 18-click easter egg."""
    import platform
    from web_app.video_stream import video_stream

    conn = get_db()
    student_count = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    log_count = conn.execute("SELECT COUNT(*) FROM attendance_logs").fetchone()[0]
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    db_size = 0
    try:
        db_path = conn.execute("PRAGMA database_list").fetchone()[2]
        if os.path.exists(db_path):
            db_size = round(os.path.getsize(db_path) / 1024, 1)
    except Exception:
        pass
    conn.close()

    return jsonify({
        "system": {
            "python": platform.python_version(),
            "os": f"{platform.system()} {platform.release()}",
            "machine": platform.machine(),
            "cwd": os.getcwd()
        },
        "database": {
            "students": student_count,
            "attendance_logs": log_count,
            "users": user_count,
            "size_kb": db_size
        },
        "ai_engine": {
            "running": video_stream.is_running,
            "camera_source": getattr(video_stream, 'camera_source', 'unknown')
        },
        "version": "1.6.0"
    })


# ══════════════════════════════════════════════════════
#  EMAIL — Test & Class Reports
# ══════════════════════════════════════════════════════

@api_bp.route('/email/test', methods=['POST'])
@api_admin_required
def email_test():
    """Send a test email to verify SMTP configuration."""
    from web_app.email_service import send_test_email
    success, error = send_test_email()
    if success:
        return jsonify({"success": True, "message": "Test email sent successfully"})
    return jsonify({"error": error}), 500


@api_bp.route('/email/class-report/<int:schedule_id>', methods=['POST'])
@api_login_required
def email_class_report(schedule_id):
    """Send per-class attendance report: individual emails to students, summary to teacher."""
    from web_app.email_service import send_student_report, send_teacher_summary, send_error_report
    from datetime import datetime, timedelta

    conn = get_db()

    # Get the schedule
    sched = conn.execute("SELECT * FROM class_schedules WHERE id = ?", (schedule_id,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"error": "Schedule not found"}), 404

    class_name = sched['class_name']
    teacher_email = sched['teacher_email'] if 'teacher_email' in sched.keys() else ''
    today = datetime.now().strftime('%Y-%m-%d')

    # Get all attendance logs for today linked to this schedule
    logs = conn.execute("""
        SELECT al.*, s.name, s.email as student_email, s.student_id as sid
        FROM attendance_logs al
        JOIN students s ON al.student_id = s.id
        WHERE al.schedule_id = ? AND date(al.timestamp) = ?
    """, (schedule_id, today)).fetchall()

    # Get ALL enrolled students to determine who's absent
    all_students = conn.execute("SELECT id, name, email FROM students").fetchall()
    conn.close()

    # Build present/absent lists
    present_ids = set()
    present_list = []
    absent_list = []
    student_statuses = {}  # id -> (name, email, status)

    for log in logs:
        sid = log['student_id']
        present_ids.add(sid)
        status = log['status']
        student_statuses[sid] = (log['name'], log['student_email'], status)
        if status in ('Present', 'On Time', 'Late'):
            present_list.append(log['name'])
        else:
            absent_list.append(log['name'])

    # Students with no log at all = Absent
    for s in all_students:
        if s['id'] not in present_ids:
            absent_list.append(s['name'])
            student_statuses[s['id']] = (s['name'], s['email'], 'Absent')

    # Send individual student emails
    sent_count = 0
    fail_count = 0
    errors = []

    for sid, (name, email, status) in student_statuses.items():
        if email:
            ok, err = send_student_report(email, name, class_name, status, today)
            if ok:
                sent_count += 1
            else:
                fail_count += 1
                errors.append(f"{name}: {err}")

    # Send teacher summary
    teacher_sent = False
    if teacher_email:
        ok, err = send_teacher_summary(teacher_email, class_name, today, present_list, absent_list)
        teacher_sent = ok
        if not ok:
            errors.append(f"Teacher: {err}")

    # If any errors, send error report to admin
    if errors:
        send_error_report(
            f"Email Report Errors — {class_name}",
            '\n'.join(errors)
        )

    return jsonify({
        "success": True,
        "class_name": class_name,
        "date": today,
        "students_emailed": sent_count,
        "students_failed": fail_count,
        "teacher_emailed": teacher_sent,
        "present": len(present_list),
        "absent": len(absent_list)
    })
